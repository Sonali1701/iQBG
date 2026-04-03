from openpyxl import Workbook
import re
from pathlib import Path

XLSX_HEADERS = [
    "Display Order*","QBG Subject Id","QBG Chapter Id","QBG Topic Id","QBG SubTopic Id",
    "Question Text","Question Image","Option Text","Option Image","Answer*",
    "Sol Text","Sol Image","Sol Video","Question Type*","Positive Marks*",
    "Negative Marks","Partial Marks","Difficulty level","Language*","Comp Id",
    "Comp Image","Section Image","QBG Question id","QR Link","PYQ Years"
]

def build_rows(q_files, s_files, language_label: str, include_tagging=False, master_tags=None):
    if master_tags is None: master_tags = {}
    q_map = {f['q_num']: f['filename'] for f in q_files} if q_files else {}
    s_map = {f['q_num']: f['filename'] for f in s_files} if s_files else {}
    
    all_nums = sorted(set(q_map.keys()) | set(s_map.keys()))
    # DIAGNOSTIC: log what we have
    if master_tags:
        print(f"[XLSX_UTILS] master_tags has {len(master_tags)} entries. Keys sample: {list(master_tags.keys())[:5]}, types: {[type(k).__name__ for k in list(master_tags.keys())[:3]]}")
        print(f"[XLSX_UTILS] all_nums sample: {all_nums[:5]}, types: {[type(n).__name__ for n in all_nums[:3]]}")
        first_key = list(master_tags.keys())[0]
        print(f"[XLSX_UTILS] sample tag_info: {master_tags[first_key]}")
    else:
        print("[XLSX_UTILS] WARNING: master_tags is EMPTY!")
    rows = []
    for n in all_nums:
        qbg_id = f"QH{n}" if language_label == "Hindi" else f"Q{n}"
        
        row = [""] * 25
        row[0] = str(n)               # Display Order
        row[6] = q_map.get(n, "")     # Question Image
        row[7] = "1~2~3~4"            # Option Text
        row[11] = s_map.get(n, "")    # Sol Image
        row[13] = "Single"            # Question Type*
        row[14] = "4"                 # Positive Marks*
        row[15] = "1"                 # Negative Marks*
        row[18] = language_label      # Language*
        row[22] = qbg_id              # QBG Question id
        
        if include_tagging:
            # Try both int and str keys to handle type mismatch
            tag_info = master_tags.get(n) or master_tags.get(str(n), {})
            # Read pre-enriched fields from tagger.py directly (no second lookup needed)
            row[1] = tag_info.get("Subject", "")
            row[2] = tag_info.get("Chapter", "")
            row[3] = tag_info.get("Topic", "")
            row[4] = tag_info.get("Subtopic", "")
            row[17] = tag_info.get("Difficulty", "")
            
        rows.append(row)
    return rows

def build_tagging_xlsx(rows, out_xlsx: Path, language_label: str, include_tagging=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tagging"

    for c, h in enumerate(XLSX_HEADERS, start=1):
        ws.cell(row=1, column=c).value = h

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["L"].width = 28
    ws.column_dimensions["S"].width = 14
    ws.column_dimensions["W"].width = 18

    for r_idx, row_data in enumerate(rows, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val

    wb.save(out_xlsx)
